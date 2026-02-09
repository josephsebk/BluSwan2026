import openpyxl
import json
import re

def parse_excel(file_path):
    print(f"DEBUG: Starting parse of {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Assuming the data is in the first sheet or a sheet named "Investor View" / "Meeting List"
        # Let's try to find a sheet with "Investor" in the name, or fall back to active.
        sheet = None
        for name in wb.sheetnames:
            if "investor" in name.lower() and "view" in name.lower():
                sheet = wb[name]
                print(f"Using sheet: {name}")
                break
        
        if not sheet:
            sheet = wb.active
            print(f"Using active sheet: {sheet.title}")

        data = []
        headers = {}
        
        # Iterate rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            print("No data found in sheet")
            return

        # Find header row
        header_row_idx = 0
        for i, row in enumerate(rows):
            row_str = [str(c).lower() if c else "" for c in row]
            # Check if any cell contains "investor" and ("fund" or "rep")
            if any("investor" in c for c in row_str) and any("fund" in c for c in row_str):
                header_row_idx = i
                for j, col in enumerate(row_str):
                    if "investor (fund)" in col: headers['fund'] = j
                    elif "investor rep" in col: headers['reps'] = j
                    elif "room" in col: headers['room'] = j
                    elif "timeslot" in col: headers['time'] = j
                    elif "founder (company)" in col: headers['founder'] = j
                    elif "investor runner" in col: headers['investorRunner'] = j
                
                # If we found fund/investor, we stop looking for header row
                if 'fund' in headers:
                    break
        
        print(f"Headers found: {headers}")

        if 'fund' not in headers:
            print("Could not find 'Fund' or 'Investor' column")
            return

        assignments = []
        id_counter = 101

        for i in range(header_row_idx + 1, len(rows)):
            row = rows[i]
            if not row[headers.get('fund', 0)]: continue # Skip empty rows

            fund = str(row[headers['fund']]).strip()
            
            # Reps
            reps_raw = str(row[headers.get('reps', -1)]) if 'reps' in headers else ""
            # Split by comma or newline
            reps = [r.strip() for r in re.split(r'[,\n]', reps_raw) if r.strip()]
            
            room = str(row[headers.get('room', -1)]).strip() if 'room' in headers else ""
            time_slot = str(row[headers.get('time', -1)]).strip() if 'time' in headers else ""
            founder = str(row[headers.get('founder', -1)]).strip() if 'founder' in headers else ""
            inv_runner = str(row[headers.get('investorRunner', -1)]).strip() if 'investorRunner' in headers else ""

            # Normalize Time Slot (Quick fix for formatting)
            # Example: "2:30 PM - 3:00 PM"
            
            if fund and time_slot:
                assignments.append({
                    "id": id_counter,
                    "investorRunner": inv_runner,
                    "fund": fund,
                    "reps": reps,
                    "investorRoom": room,
                    "timeSlot": time_slot,
                    "founderCompany": founder
                })
                id_counter += 1

        print(json.dumps(assignments, indent=2))

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    parse_excel("BluSwan final.xlsx")
